"""Smart JSON / Excel export for extracted document data.

The extracted data is hierarchical (Pydantic models with nested Party,
LineItem, etc.). For human-friendly Excel we:

- Flatten nested dicts with dot notation onto a `Document` sheet
  (e.g. `seller.name`, `seller.country`).
- Give each list field its own sheet with one row per item
  (e.g. `line_items` -> sheet named "Line Items").
- Auto-fit column widths and bold the header row.

The JSON export is the raw `extracted_data` pretty-printed with
UTF-8 (so Cyrillic / Uzbek chars don't get escaped).
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ── JSON ────────────────────────────────────────────────────────────


def to_json_bytes(data: dict[str, Any]) -> bytes:
    """UTF-8 JSON bytes, pretty-printed, with non-ASCII preserved."""
    return json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")


# ── CSV ─────────────────────────────────────────────────────────────


def to_csv_bytes(data: dict[str, Any]) -> bytes:
    """Flatten extracted data into a UTF-8 CSV (BOM for Excel auto-detection).

    Scalar and primitive-list fields appear as Field/Value rows. Each list-of-
    dicts field (e.g. line_items) gets its own section with a header row.
    """
    main_rows, list_fields = _split_main_and_lists(data)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Field", "Value"])
    for row in main_rows:
        writer.writerow([row["Field"], row["Value"]])

    for key, items in list_fields.items():
        if not items:
            continue
        writer.writerow([])
        writer.writerow([f"[{key}]"])
        df = _flatten_list_to_dataframe(items)
        writer.writerow(list(df.columns))
        for _, row_data in df.iterrows():
            writer.writerow([str(v) if v != "" else "" for v in row_data])

    return output.getvalue().encode("utf-8-sig")


# ── Excel ───────────────────────────────────────────────────────────


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_MAX_COL_WIDTH = 50


def to_excel_bytes(data: dict[str, Any]) -> bytes:
    """Serialise extracted data into a multi-sheet .xlsx file."""
    main_rows, list_fields = _split_main_and_lists(data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Main sheet: flat field/value table for everything that isn't a list.
        if main_rows:
            df_main = pd.DataFrame(main_rows, columns=["Field", "Value"])
            df_main.to_excel(writer, sheet_name="Document", index=False)
            _polish_sheet(writer.sheets["Document"], df_main)
        elif not list_fields:
            # openpyxl refuses to write a workbook with zero visible sheets,
            # so for fully-empty input we emit a single placeholder sheet.
            df_empty = pd.DataFrame([{"Field": "(no extracted data)", "Value": ""}])
            df_empty.to_excel(writer, sheet_name="Document", index=False)
            _polish_sheet(writer.sheets["Document"], df_empty)

        # Each list field becomes its own sheet of rows.
        for key, items in list_fields.items():
            if not items:
                continue
            df = _flatten_list_to_dataframe(items)
            sheet_name = _excel_safe_sheet_name(key)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _polish_sheet(writer.sheets[sheet_name], df)

    return output.getvalue()


# ── Helpers ─────────────────────────────────────────────────────────


def _split_main_and_lists(
    data: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, list[Any]]]:
    """Walk the dict; return (flat rows, list-valued fields)."""
    rows: list[dict[str, Any]] = []
    lists: dict[str, list[Any]] = {}

    def visit(prefix: str, value: Any) -> None:
        if isinstance(value, dict):
            for k, v in value.items():
                visit(f"{prefix}.{k}" if prefix else k, v)
        elif isinstance(value, list):
            # Lists of primitives stay on the main sheet, joined with commas.
            if all(not isinstance(item, (dict, list)) for item in value):
                rows.append({"Field": prefix, "Value": ", ".join(str(v) for v in value)})
            else:
                lists[prefix] = value
        else:
            rows.append({"Field": prefix, "Value": _display(value)})

    for k, v in data.items():
        visit(k, v)
    return rows, lists


def _flatten_list_to_dataframe(items: list[Any]) -> pd.DataFrame:
    """Normalise a list of dicts into a flat DataFrame."""
    flat_rows: list[dict[str, Any]] = []
    for item in items:
        if isinstance(item, dict):
            flat_rows.append(_flatten_one(item))
        else:
            flat_rows.append({"value": _display(item)})
    return pd.DataFrame(flat_rows)


def _flatten_one(d: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    flat: dict[str, Any] = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            flat.update(_flatten_one(v, key))
        elif isinstance(v, list):
            flat[key] = ", ".join(_display(x) for x in v)
        else:
            flat[key] = _display(v)
    return flat


def _display(value: Any) -> Any:
    """Render None as empty cell; pass everything else unchanged."""
    return "" if value is None else value


def _excel_safe_sheet_name(name: str) -> str:
    """Excel sheet names: ≤31 chars, no `: / \\ ? * [ ]`."""
    cleaned = (
        name.replace("_", " ").replace(":", "").replace("/", "")
        .replace("\\", "").replace("?", "").replace("*", "")
        .replace("[", "").replace("]", "")
    )
    return cleaned.title()[:31]


def _polish_sheet(worksheet, df: pd.DataFrame) -> None:
    """Bold header row, auto-fit widths, freeze the top row."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Compute a width that fits the longest cell, capped.
        max_len = max(
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
            len(str(col_name)),
        )
        width = min(max_len + 2, _MAX_COL_WIDTH)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

        # Header styling
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    worksheet.freeze_panes = "A2"
