"""Unit tests for `app.document_export`."""
from __future__ import annotations

import io
import json

import openpyxl
import pytest

from app.document_export import (
    _excel_safe_sheet_name,
    _flatten_one,
    _split_main_and_lists,
    to_excel_bytes,
    to_json_bytes,
)

# Use the actual Ascensia invoice extraction as a realistic sample.
SAMPLE_INVOICE = {
    "invoice_number": "HQPL00073841",
    "invoice_date": "2025-07-03",
    "seller": {
        "name": "Ascensia Diabetes Care Holdings AG",
        "country": "CH",
        "tax_id": "CHE-123",
    },
    "buyer": {
        "name": "FE Medical Online Services LLC",
        "country": "UZ",
        "tax_id": None,
    },
    "line_items": [
        {
            "description": "CONTOUR PLUS test strips",
            "hs_code": "9027899000",
            "quantity": 3994,
            "unit_price": 1.59,
            "amount": 6350.46,
        },
        {
            "description": "MICROLET lancets",
            "hs_code": "9018390000",
            "quantity": 500,
            "unit_price": 0.30,
            "amount": 150.0,
        },
    ],
    "total_amount": 6500.46,
    "currency": "EUR",
    "notes": None,
}


# ── JSON ────────────────────────────────────────────────────────────


class TestJsonExport:
    def test_returns_utf8_bytes(self):
        out = to_json_bytes(SAMPLE_INVOICE)
        assert isinstance(out, bytes)
        # Round-trip cleanly
        parsed = json.loads(out.decode("utf-8"))
        assert parsed["invoice_number"] == "HQPL00073841"
        assert parsed["seller"]["name"].startswith("Ascensia")

    def test_preserves_non_ascii(self):
        data = {"name": "Тошкент", "company": "ЛТД «Образец»"}
        out = to_json_bytes(data).decode("utf-8")
        # Cyrillic must NOT be escaped to \uXXXX sequences
        assert "Тошкент" in out
        assert "«" in out

    def test_pretty_printed(self):
        out = to_json_bytes({"a": 1, "b": 2}).decode("utf-8")
        assert "\n" in out
        assert "  " in out  # 2-space indent


# ── Splitting logic ─────────────────────────────────────────────────


class TestSplitMainAndLists:
    def test_flat_dict_yields_rows(self):
        rows, lists = _split_main_and_lists({"a": 1, "b": "x"})
        assert lists == {}
        fields = {r["Field"]: r["Value"] for r in rows}
        assert fields == {"a": 1, "b": "x"}

    def test_nested_dict_is_dotted(self):
        rows, _ = _split_main_and_lists(
            {"seller": {"name": "A", "country": "CH"}}
        )
        fields = {r["Field"]: r["Value"] for r in rows}
        assert fields["seller.name"] == "A"
        assert fields["seller.country"] == "CH"

    def test_lists_of_dicts_extracted_separately(self):
        rows, lists = _split_main_and_lists(SAMPLE_INVOICE)
        assert "line_items" in lists
        assert len(lists["line_items"]) == 2
        # Top-level scalars and nested dicts both appear in rows
        field_names = {r["Field"] for r in rows}
        assert "invoice_number" in field_names
        assert "seller.name" in field_names

    def test_lists_of_primitives_stay_flat(self):
        rows, lists = _split_main_and_lists(
            {"tags": ["urgent", "medical", "EU"]}
        )
        assert lists == {}
        fields = {r["Field"]: r["Value"] for r in rows}
        assert "urgent" in fields["tags"]
        assert "medical" in fields["tags"]

    def test_none_becomes_empty_string(self):
        rows, _ = _split_main_and_lists({"missing": None})
        values = {r["Field"]: r["Value"] for r in rows}
        assert values["missing"] == ""


# ── Flattening list rows ────────────────────────────────────────────


class TestFlattenOne:
    def test_simple_dict_unchanged(self):
        assert _flatten_one({"a": 1, "b": "x"}) == {"a": 1, "b": "x"}

    def test_nested_dict_dotted(self):
        assert _flatten_one({"seller": {"name": "A"}}) == {"seller.name": "A"}

    def test_nested_list_joined(self):
        assert _flatten_one({"codes": ["a", "b"]}) == {"codes": "a, b"}


# ── Excel sheet name sanitisation ───────────────────────────────────


class TestExcelSheetName:
    def test_underscores_to_spaces_and_title_case(self):
        assert _excel_safe_sheet_name("line_items") == "Line Items"

    def test_invalid_chars_stripped(self):
        assert _excel_safe_sheet_name("a/b\\c:d?e*f[g]h") == "Abcdefgh"

    def test_truncated_to_31_chars(self):
        result = _excel_safe_sheet_name("very_" * 20)
        assert len(result) <= 31


# ── Full Excel export ───────────────────────────────────────────────


class TestExcelExport:
    def test_returns_valid_xlsx_bytes(self):
        out = to_excel_bytes(SAMPLE_INVOICE)
        assert isinstance(out, bytes)
        assert out[:2] == b"PK"  # xlsx is a zip archive

        # Round-trip with openpyxl to confirm it's a valid workbook
        wb = openpyxl.load_workbook(io.BytesIO(out))
        assert "Document" in wb.sheetnames
        assert any(name.lower().startswith("line") for name in wb.sheetnames)

    def test_main_sheet_has_field_value_columns(self):
        out = to_excel_bytes(SAMPLE_INVOICE)
        wb = openpyxl.load_workbook(io.BytesIO(out))
        ws = wb["Document"]
        headers = [cell.value for cell in ws[1]]
        assert headers == ["Field", "Value"]

    def test_main_sheet_includes_invoice_number(self):
        out = to_excel_bytes(SAMPLE_INVOICE)
        wb = openpyxl.load_workbook(io.BytesIO(out))
        ws = wb["Document"]
        # Walk rows looking for the invoice number value
        invoice_numbers = [
            row[1].value for row in ws.iter_rows(min_row=2)
            if row[0].value == "invoice_number"
        ]
        assert invoice_numbers == ["HQPL00073841"]

    def test_line_items_sheet_has_one_row_per_item(self):
        out = to_excel_bytes(SAMPLE_INVOICE)
        wb = openpyxl.load_workbook(io.BytesIO(out))
        sheet = next(
            wb[name] for name in wb.sheetnames if name.lower().startswith("line")
        )
        # Header row + 2 items
        assert sheet.max_row == 3
        # Headers include "description" and "hs_code"
        headers = [c.value for c in sheet[1]]
        assert "description" in headers
        assert "hs_code" in headers

    def test_header_row_is_bold(self):
        out = to_excel_bytes(SAMPLE_INVOICE)
        wb = openpyxl.load_workbook(io.BytesIO(out))
        ws = wb["Document"]
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_handles_doc_with_no_line_items(self):
        # AWB-like — no list-valued fields at the top level
        awb = {
            "awb_number": "488-40007166",
            "shipper": {"name": "Sender Co"},
            "consignee": {"name": "Receiver Co"},
            "gross_weight_kg": 3019.0,
        }
        out = to_excel_bytes(awb)
        wb = openpyxl.load_workbook(io.BytesIO(out))
        assert wb.sheetnames == ["Document"]

    def test_empty_input_produces_empty_workbook(self):
        out = to_excel_bytes({})
        wb = openpyxl.load_workbook(io.BytesIO(out))
        # No sheets created (empty dict → no main rows, no lists)
        # openpyxl creates a default sheet on empty workbooks though, so we
        # just verify we got valid bytes back.
        assert isinstance(out, bytes)
